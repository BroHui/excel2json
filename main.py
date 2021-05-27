# coding: utf-8

import logging
import warnings

import yaml
import os
import re
import string
from openpyxl import load_workbook
import xlrd

logging.basicConfig(level=logging.DEBUG)

BLANK_LINES = 3
EXCEL_XLS = 2003
EXCEL_XLSX = 2007

TABLE_STATIC = 1
TABLE_FLOATING = 2


class UnknownFiletype(Exception):
    pass


class ExcelDriver(object):
    """
        抽象出Excel操作对象，兼容97/07版格式
    """
    def __init__(self, excel_file_name=""):
        self.excel_type = 0
        self.file_name = excel_file_name
        if excel_file_name.endswith('.xls'):
            self.excel_type = EXCEL_XLS
        elif excel_file_name.endswith('.xlsx'):
            self.excel_type = EXCEL_XLSX
        else:
            warnings.warn("Unknown file type, {}".format(excel_file_name))
            raise UnknownFiletype()
        self.wb = None  # work book
        self.ws = None  # work sheet
        self.load_file(excel_file_name)

    def load_file(self, filename):
        if not filename:
            logging.warning("Missing excel file name.")
            return
        if self.excel_type == EXCEL_XLSX:
            self.wb = load_workbook(filename)
            self.sheets = self.wb.sheetnames
            self.ws = self.wb.active
        elif self.excel_type == EXCEL_XLS:
            self.wb = xlrd.open_workbook(filename)
            self.ws = self.wb.sheet_by_index(0)  # 默认第一页sheet
        return

    def change_sheet(self, sheet_index=0):
        if self.excel_type == EXCEL_XLSX:
            self.wb.active = sheet_index
            self.ws = self.wb.active
        elif self.excel_type == EXCEL_XLS:
            self.ws = self.wb.sheet_by_index(sheet_index)

    def cell_name_to_number(self, cell_name):
        match = re.match(r"([a-z]+)([0-9]+)", cell_name, re.I)
        if match:
            colx, rowx = match.groups()
            colx = ord(colx)
            if colx <= 65:
                colx = 0
            else:
                colx = colx - 65
            rowx = int(rowx)
            return colx, rowx
        else:
            return 0, 0

    def get_shift_cell_name(self, std_cell, shift_list):
        if not shift_list or len(shift_list) != 4:
            return std_cell
        cell_colx, cell_rowx = self.cell_name_to_number(std_cell)
        if shift_list[0]:  # 上偏移
            cell_rowx -= shift_list[0]
        elif shift_list[1]:  # 下偏移
            cell_rowx += shift_list[1]
        elif shift_list[2]:  # 左偏移
            cell_colx -= shift_list[2]
        elif shift_list[3]:  # 右偏移
            cell_colx += shift_list[3]
        if self.excel_type == EXCEL_XLS:
            return xlrd.cellname(cell_rowx, cell_colx)
        elif self.excel_type == EXCEL_XLSX:
            cell_name = "{}{}".format(chr(cell_colx+65), cell_rowx)
            return cell_name

    def get_cell_value(self, cell_pos):
        if self.excel_type == EXCEL_XLSX:
            return self.ws[cell_pos].value
        elif self.excel_type == EXCEL_XLS:
            colx, rowx = self.cell_name_to_number(cell_pos)
            try:
                return self.ws.cell_value(rowx, colx)
            except IndexError:
                return ''


class ExcelBook(object):
    def __init__(self):
        self.parse_mode = 1
        self.excel = None
        self.cols = []
        self.row_range = []
        self.headers = []
        self.table_type = TABLE_STATIC
        self.pick_list = {}
        self.desc_shift = []

    def load_yaml(self, yaml_filename=""):
        if not yaml_filename:
            return {}
        data = {}
        if os.path.exists(yaml_filename):
            logging.info("find {}, read config".format(yaml_filename))
            with open(yaml_filename, encoding='UTF-8') as fp:
                data = yaml.load(fp, yaml.FullLoader)
                logging.debug("the YAML config is {}".format(data))
        return data

    def load(self, xlsx_file_name="", yaml_config_file="", yaml_config={}):
        if not xlsx_file_name:
            logging.warning("Missing xlsx file name.")
            return
        self.excel = ExcelDriver(xlsx_file_name)

        # Load desc yaml config
        if not yaml_config_file:
            yaml_filename = xlsx_file_name.split('.')[0] + '.yml'
        else:
            yaml_filename = yaml_config_file
        skip_level, skip_table_headers = 0, 1
        if yaml_config:
            data = yaml_config
            logging.info("Using yaml_config instead. {}".format(yaml_config))
        else:
            data = self.load_yaml(yaml_filename)
        logging.debug(data)
        if 'static' in data.keys():
            headers = data.get('static', {}).get('headers', {})
            skip_level = headers.get('skip_level', 0)
            skip_table_headers = headers.get('total_high', 1)
            self.table_type = TABLE_STATIC
        elif 'floating' in data.keys():
            self.table_type = TABLE_FLOATING

        if self.table_type == TABLE_STATIC:
            # peek the rows and cols at the excel
            _, _, self.cols = self.get_cols_range(skip_level=skip_level)
            row_start, row_end, _ = self.get_rows_range(skip_table_headers=skip_table_headers)
            self.row_range = (row_start, row_end)

            # try to load headers
            headers_row = skip_table_headers - 1 if self.excel.excel_type == EXCEL_XLS else skip_table_headers
            self.headers = self.get_headers(headers_row=headers_row)
        elif self.table_type == TABLE_FLOATING:
            floating_conf = data['floating']
            self.pick_list = floating_conf['values']
            if 'desc_shift' in floating_conf:
                self.desc_shift = floating_conf['desc_shift']

    def get_cols_range(self, skip_level=0):
        """
            excel cols range peek
        :return:
        """
        logging.debug("get cols range")
        col_start = string.ascii_uppercase[0]
        col_end = ''
        col_array = []
        blank_col_count = 0
        for alphabet in string.ascii_uppercase:
            cell_pos = '{}{}'.format(alphabet, 1+skip_level)
            if self.excel.get_cell_value(cell_pos):
                blank_col_count = 0
                col_array.append(alphabet)
                col_end = alphabet
                continue
            # blank col, start counting
            blank_col_count += 1
            if blank_col_count == 3:
                break
        logging.debug("col_start {}, col_end {}, col_array {}".format(col_start, col_end, col_array))
        return col_start, col_end, col_array

    def get_rows_range(self, skip_table_headers=1, the_first_col='A'):
        """
            excel rows range peek
            default table header rows set as one.
        :return:
        """
        logging.debug("getting rows range from excel, skip table headers {}, and the first col is {}".format(
            skip_table_headers, the_first_col
        ))
        row_start = 1 + skip_table_headers
        row_end = 1 + skip_table_headers
        row_array = []
        blank_row_count = 0
        while 1:
            cell_pos = '{}{}'.format(the_first_col, row_end)
            if self.excel.get_cell_value(cell_pos):
                blank_row_count = 0
                row_array.append(cell_pos)
                row_end += 1
                continue
            # blank row, start counting
            blank_row_count += 1
            if blank_row_count == 3:
                row_end -= 1
                break
        logging.debug("row start {}, row end {}".format(row_start, row_end))
        return row_start, row_end, row_array

    def get_headers(self, headers_row=1, headers_blocks=[]):
        headers = []
        # _, _, col_range = self.get_cols_range()
        for col in self.cols:
            header_pos = "{}{}".format(col, headers_row)
            cell_value = self.excel.get_cell_value(header_pos)
            if not cell_value:
                continue
            headers.append(cell_value)
        logging.info("this excel file headers {}".format(headers))
        return headers

    def get_data(self):
        if not self.excel:
            logging.error("Excel file object not init.")
            return
        data = []
        if self.table_type == TABLE_STATIC:
            row_start, row_end = self.row_range
            for i in range(row_start, row_end+1):
                single_data = {}
                for k in self.cols:
                    pos = "{}{}".format(k, i)
                    cell_value = self.excel.get_cell_value(pos) or ""
                    t_h = self.headers[self.cols.index(k)]
                    single_data[t_h] = cell_value
                data.append(single_data)
        elif self.table_type == TABLE_FLOATING:
            for pick_cell in self.pick_list:
                cell_key = self.pick_list[pick_cell]
                cell_value = self.excel.get_cell_value(pick_cell) or ""
                single_data = {
                    'key': cell_key,
                    'value': cell_value
                }
                # 如果有注释
                if self.desc_shift:
                    desc_cell = self.excel.get_shift_cell_name(pick_cell, self.desc_shift)
                    desc_value = self.excel.get_cell_value(desc_cell) or ""
                    single_data['desc'] = desc_value
                data.append(single_data)
        return data


class SimpleExcelBook(ExcelBook):
    def __init__(self):
        super(SimpleExcelBook, self).__init__()


if __name__ == '__main__':
    book = SimpleExcelBook()
    # book.load('demo1.xlsx')
    book.load('demo2.xlsx', yaml_config={'static': {'headers': {'skip_level': 1, 'total_high': 2}}})
    # book.load('demo3.xlsx')
    # book.load('sample2.xls')
    data = book.get_data()
    print(data)

